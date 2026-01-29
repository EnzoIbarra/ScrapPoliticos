
import json
import pandas as pd
from pathlib import Path

def export_to_excel():
    data_dir = Path("data")
    valid_file = data_dir / "valid_results.json"
    retry_file = data_dir / "retry_queue.json"
    output_file = data_dir / "scrapped_data.xlsx"

    # 1. Procesar Resultados Válidos
    valid_rows = []
    if valid_file.exists():
        with open(valid_file, 'r', encoding='utf-8') as f:
            valid_data = json.load(f)
            
        for item in valid_data:
            muni = item.get("municipality")
            url = item.get("url")
            people = item.get("data", [])
            for person in people:
                row = {
                    "Municipio": muni,
                    "URL Municipio": url,
                    "Nombre": person.get("nombre", "(Sin contactos encontrados)"),
                    "Cargo": person.get("cargo"),
                    "Partido": person.get("partido"),
                    "Email": person.get("email"),
                    "Tipo": person.get("tipo"),
                    "Source URL": person.get("source_url")
                }
                valid_rows.append(row)
    
    df_valid = pd.DataFrame(valid_rows)

    # 2. Procesar Fallidos/Pendientes
    retry_rows = []
    if retry_file.exists():
        with open(retry_file, 'r', encoding='utf-8') as f:
            retry_data = json.load(f)
            
        for item in retry_data:
            retry_rows.append({
                "Municipio": item.get("municipality"),
                "URL Municipio": item.get("url"),
                "Nombre": "FALLO DE CONEXIÓN",  # Indicador claro
                "Cargo": None,
                "Partido": None,
                "Email": None,
                "Tipo": None,
                "Source URL": None
            })
            
    df_retry = pd.DataFrame(retry_rows)

    # 3. Guardar en Excel con formato consolidado
    print(f"📊 Generando Excel Consolidado con {len(valid_rows) + len(retry_rows)} registros...")
    
    # Unificar todo en un solo DataFrame
    df_final = pd.concat([df_valid, df_retry], ignore_index=True)
    
    # Ordenar por Municipio
    if not df_final.empty:
        df_final.sort_values(by="Municipio", inplace=True)
    
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df_final.to_excel(writer, sheet_name='Consolidado_Municipios', index=False)
            
            worksheet = writer.sheets['Consolidado_Municipios']
            
            # Estilos Corporativos
            header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
            header_fill = PatternFill("solid", fgColor="003366") # Azul oscuro
            alignment_center = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'), 
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            # 1. Formatear Cabeceras
            for col_num, value in enumerate(df_final.columns.values):
                cell = worksheet.cell(row=1, column=col_num+1)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            # 2. Formatear Datos y Ajustar Ancho
            for col_idx, column_cells in enumerate(worksheet.columns, 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)

                # Header len
                header_val = str(column_cells[0].value)
                max_length = len(header_val) + 4

                for cell in column_cells[1:]:
                    cell.border = thin_border
                    cell.alignment = alignment_center
                    
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('http'):
                        cell.font = Font(color="0563C1", underline="single")
                    
                    if cell.value:
                        item_len = len(str(cell.value))
                        if item_len > max_length:
                            max_length = item_len
                
                # Topar ancho
                worksheet.column_dimensions[column_letter].width = min(max_length, 60)

            # 3. Congelar paneles y Filtros
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions

        print(f"✅ Exportación completada con formato corporativo: {output_file}")
    except ImportError:
        print("❌ Error: Falta librería 'openpyxl'. Instalando...")
        import subprocess
        subprocess.check_call(["pip", "install", "openpyxl"])
        # Reintentar una vez
        export_to_excel()
    except Exception as e:
        print(f"❌ Error exportando Excel: {e}")

if __name__ == "__main__":
    export_to_excel()
